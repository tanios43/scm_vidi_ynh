#!/usr/bin/env python3
"""
SCM Vidi — Export Excel annuel
Usage : python export_excel.py [scm_vidi_data.json] [annee]
        python export_excel.py scm_vidi_2025.json 2025
"""

import json, sys, math
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── Helpers styles ─────────────────────────────────────────────────────────────
FONT = "Arial"
MONEY = '#,##0.00 €'
PCT   = '0.0%'

C = {
    'dk':    '1A2744', 'md':    '2E5090', 'lt':    'BDD7EE',
    'ys':    'D6ECD5', 'ys2':   'A8D5A6', 'ys3':   '2A6B28',
    'py':    'FDF0CC', 'py2':   'F7D97A', 'py3':   '7A5A00',
    'gn':    'FAE0D4', 'gn2':   'F3B49A', 'gn3':   '8B2500',
    'wh':    'FFFFFF', 'bg':    'F7F5F0', 'tot':   'EEF2FF',
    'rev':   'FDF5E0', 'av':    'EDE0F5',
}

def thin():
    s = Side(style='thin')
    return Border(top=s, bottom=s, left=s, right=s)

def medium():
    m = Side(style='medium')
    return Border(top=m, bottom=m, left=m, right=m)

def hc(ws, row, col, val='', bg=C['dk'], fg=C['wh'], bold=True, sz=10,
       align='center', wrap=False, fmt=None, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, bold=bold, color=fg, size=sz, italic=italic)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = thin()
    if fmt: c.number_format = fmt
    return c

def dc(ws, row, col, val=None, bg=C['wh'], color='000000', bold=False,
       fmt=None, align='center', italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, color=color, bold=bold, size=10, italic=italic)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = thin()
    if fmt: c.number_format = fmt
    return c

def merge_hc(ws, r1, c1, r2, c2, val='', bg=C['dk'], fg=C['wh'], bold=True, sz=10, align='center'):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=val)
    c.font = Font(name=FONT, bold=bold, color=fg, size=sz)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = thin()
    return c

MOIS = ["Janvier","Février","Mars","Avril","Mai","Juin",
        "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

# ── Chargement des données ─────────────────────────────────────────────────────
def load(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def get_effectif(data, yr, mo, cid):
    reel = (data.get('montantsReels', {})
                .get(str(yr), {}).get(str(mo), {}).get(str(cid)))
    if reel is not None:
        return reel
    ch = next((c for c in data['charges'] if c['id'] == cid), None)
    return ch['montant'] if ch else 0

def dj_count(data, yr, mo, cab, aid):
    slots = (data.get('plannings', {})
                 .get(str(yr), {}).get(str(mo), {}).get(cab, {}))
    return sum(1 for v in slots.values() if v == aid)

def dj_total_cab(data, yr, mo, cab, assocs):
    return sum(dj_count(data, yr, mo, cab, a['id']) for a in assocs if a.get('actif'))

def get_ca(data, yr, mo, aid):
    return (data.get('caPuy', {})
                .get(str(yr), {}).get(str(mo), {}).get(str(aid), 0))

def get_avance(data, yr, mo, aid):
    return (data.get('avances', {})
                .get(str(yr), {}).get(str(mo), {}).get(str(aid), 0))

def taux_redev(data):
    return data.get('tauxRedev', 0.10)

def calc_repartition(data, yr, mo):
    """Retourne dict aid -> {y, p, g, rev, tot, av, reste}"""
    assocs = [a for a in data['associes'] if a.get('actif')]
    nb = len(assocs)
    charges = data['charges']
    tx = taux_redev(data)

    tY = dj_total_cab(data, yr, mo, 'Yssingeaux', assocs)
    tP = dj_total_cab(data, yr, mo, 'Le Puy', assocs)

    res = {a['id']: {'y': 0, 'p': 0, 'g': 0, 'rev': 0} for a in assocs}
    for ch in charges:
        m = get_effectif(data, yr, mo, ch['id'])
        if ch['type'] == 'Yssingeaux':
            for a in assocs:
                dj = dj_count(data, yr, mo, 'Yssingeaux', a['id'])
                res[a['id']]['y'] += (dj / tY * m) if tY > 0 else 0
        elif ch['type'] == 'Le Puy':
            for a in assocs:
                dj = dj_count(data, yr, mo, 'Le Puy', a['id'])
                res[a['id']]['p'] += (dj / tP * m) if tP > 0 else 0
        else:
            for a in assocs:
                res[a['id']]['g'] += (m / nb) if nb > 0 else 0

    for a in assocs:
        ca = get_ca(data, yr, mo, a['id'])
        res[a['id']]['rev'] = ca * tx
        tot = sum(res[a['id']].values())
        av = get_avance(data, yr, mo, a['id'])
        res[a['id']]['tot'] = tot
        res[a['id']]['av'] = av
        res[a['id']]['reste'] = max(0, tot - av)

    return res

# ── ONGLET RÉCAP ANNUEL ────────────────────────────────────────────────────────
def build_recap(wb, data, yr):
    ws = wb.active
    ws.title = f"Récap {yr}"
    ws.sheet_view.showGridLines = False

    assocs = [a for a in data['associes'] if a.get('actif')]
    NA = len(assocs)

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 28
    for i in range(3, 16 + NA * 4):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Titre
    ws.row_dimensions[1].height = 32
    nb_cols = 2 + 12 + 1
    merge_hc(ws, 1, 1, 1, nb_cols,
             f'SCM Vidi — Récapitulatif annuel {yr}', sz=14)

    # Sous-titre
    ws.row_dimensions[2].height = 18
    merge_hc(ws, 2, 1, 2, nb_cols,
             f'Généré le {date.today().strftime("%d/%m/%Y")} · Taux redevance matériel : {taux_redev(data)*100:.1f}%',
             bg=C['md'], sz=9, bold=False)

    # En-têtes mois
    ws.row_dimensions[4].height = 28
    hc(ws, 4, 1, '', bg=C['bg'])
    hc(ws, 4, 2, 'Orthoptiste / Indicateur', bg=C['lt'], fg=C['dk'], align='left')
    for mi, m in enumerate(MOIS):
        hc(ws, 4, 3 + mi, m, bg=C['lt'], fg=C['dk'], wrap=True)
    hc(ws, 4, 15, 'TOTAL ANNUEL', bg=C['dk'])

    row = 5
    grand_totals = [0] * 13  # 12 mois + total

    for ai, a in enumerate(assocs):
        bg_a = C['bg'] if ai % 2 == 0 else C['wh']
        # Ligne nom
        ws.row_dimensions[row].height = 20
        hc(ws, row, 1, ai + 1, bg=C['dk'], sz=9)
        hc(ws, row, 2, a['nom'], bg=C['dk'], align='left', sz=11)
        for ci in range(3, 16):
            hc(ws, row, ci, '', bg=C['dk'])
        row += 1

        labels = [
            ('Part Yssingeaux',    'y',     C['ys'],  C['ys3']),
            ('Part Le Puy',        'p',     C['py'],  C['py3']),
            ('Part Général',       'g',     C['gn'],  C['gn3']),
            ('Redevance matériel', 'rev',   C['rev'], '7A5A00'),
            ('Total dû',           'tot',   C['tot'], C['md']),
            ('Avance versée',      'av',    C['av'],  '4A235A'),
            ('Reste à payer',      'reste', C['dk'],  C['wh']),
        ]

        annual = {k: 0 for k in ['y', 'p', 'g', 'rev', 'tot', 'av', 'reste']}

        for lbl, key, bg, fg in labels:
            ws.row_dimensions[row].height = 16
            dc(ws, row, 1, '', bg=bg)
            dc(ws, row, 2, lbl, bg=bg, color=fg, align='left',
               bold=(key in ['tot', 'reste']))
            mo_vals = []
            for mo in range(12):
                rep = calc_repartition(data, yr, mo)
                val = rep.get(a['id'], {}).get(key, 0)
                mo_vals.append(val)
                annual[key] += val
                c = dc(ws, row, 3 + mo, val, bg=bg, color=fg,
                       fmt=MONEY, bold=(key in ['tot', 'reste']))
                if key == 'reste' and val < 0.01:
                    c.font = Font(name=FONT, color='1D9E75', bold=True, size=10)
            # Total annuel
            dc(ws, row, 15, annual[key], bg=bg, color=fg,
               fmt=MONEY, bold=(key in ['tot', 'reste']))
            row += 1

        # Ligne vide séparatrice
        for ci in range(1, 16):
            ws.cell(row=row, column=ci).fill = PatternFill('solid', fgColor='F0EDE8')
        row += 1

    # Ligne total SCM
    ws.row_dimensions[row].height = 22
    merge_hc(ws, row, 1, row, 2, 'TOTAL CHARGES SCM', bg=C['dk'], sz=11)
    for mo in range(12):
        tot_mo = sum(
            calc_repartition(data, yr, mo).get(a['id'], {}).get('tot', 0)
            for a in assocs)
        dc(ws, row, 3 + mo, tot_mo, bg=C['dk'], color=C['wh'],
           fmt=MONEY, bold=True)
    all_mo = [
        sum(calc_repartition(data, yr, mo).get(a['id'], {}).get('tot', 0)
            for a in assocs)
        for mo in range(12)]
    dc(ws, row, 15, sum(all_mo), bg=C['dk'], color=C['wh'],
       fmt=MONEY, bold=True)

    ws.freeze_panes = 'C5'
    return ws

# ── ONGLET MENSUEL ─────────────────────────────────────────────────────────────
def build_mois(wb, data, yr, mo):
    ws = wb.create_sheet(MOIS[mo][:4] + '.')
    ws.sheet_view.showGridLines = False
    assocs = [a for a in data['associes'] if a.get('actif')]
    charges = data['charges']
    NA = len(assocs)
    tx = taux_redev(data)

    # Largeurs
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 4
    for i in range(3, 3 + NA + 2):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Titre
    ws.row_dimensions[1].height = 30
    nc = 3 + NA
    merge_hc(ws, 1, 1, 1, nc,
             f'SCM Vidi — {MOIS[mo]} {yr}', sz=13)

    row = 3
    moisSuiv = (mo + 1) % 12
    anSuiv = yr + (1 if mo == 11 else 0)
    merge_hc(ws, row, 1, row, nc,
             f'Redevances payables le 1er {MOIS[moisSuiv]} {anSuiv}',
             bg=C['md'], sz=9, bold=False)
    row += 2

    # ── Section 1 : PLANNING DJ ────────────────────────────────────────────────
    merge_hc(ws, row, 1, row, nc, '1. DEMI-JOURNÉES TRAVAILLÉES', bg=C['md'], sz=10)
    row += 1
    hc(ws, row, 1, 'Cabinet', bg=C['lt'], fg=C['dk'], align='left')
    hc(ws, row, 2, '', bg=C['lt'])
    for i, a in enumerate(assocs):
        hc(ws, row, 3 + i, a['nom'], bg=C['lt'], fg=C['dk'], wrap=True)
    hc(ws, row, 3 + NA, 'TOTAL', bg=C['dk'])
    row += 1

    for cab, bg in [('Yssingeaux', C['ys']), ('Le Puy', C['py'])]:
        tot_cab = dj_total_cab(data, yr, mo, cab, assocs)
        dc(ws, row, 1, cab, bg=bg, align='left', bold=True)
        dc(ws, row, 2, '', bg=bg)
        for i, a in enumerate(assocs):
            dj = dj_count(data, yr, mo, cab, a['id'])
            dc(ws, row, 3 + i, dj, bg=bg,
               color=C['ys3'] if cab == 'Yssingeaux' else C['py3'],
               bold=True)
        dc(ws, row, 3 + NA, tot_cab, bg=bg, bold=True)
        row += 1
    row += 1

    # ── Section 2 : CHARGES ────────────────────────────────────────────────────
    merge_hc(ws, row, 1, row, nc, '2. CHARGES DU MOIS', bg=C['md'], sz=10)
    row += 1
    hc(ws, row, 1, 'Libellé', bg=C['lt'], fg=C['dk'], align='left')
    hc(ws, row, 2, 'Type', bg=C['lt'], fg=C['dk'])
    hc(ws, row, 3, 'Montant habituel', bg=C['lt'], fg=C['dk'], wrap=True)
    hc(ws, row, 4, f'Montant réel {MOIS[mo][:4]}.', bg=C['lt'], fg=C['dk'], wrap=True)
    for ci in range(5, nc + 1):
        hc(ws, row, ci, '', bg=C['lt'])
    ws.row_dimensions[row].height = 30
    row += 1

    type_bg = {'Yssingeaux': C['ys'], 'Le Puy': C['py'], 'Général': C['gn']}
    type_tot = {'Yssingeaux': 0, 'Le Puy': 0, 'Général': 0}
    tot_hab = 0; tot_reel = 0

    for ch in charges:
        bg = type_bg.get(ch['type'], C['wh'])
        hab = ch['montant']
        reel = get_effectif(data, yr, mo, ch['id'])
        type_tot[ch['type']] += reel
        tot_hab += hab; tot_reel += reel
        dc(ws, row, 1, ch['lib'], bg=bg, align='left')
        dc(ws, row, 2, ch['type'], bg=bg)
        dc(ws, row, 3, hab, bg=bg, fmt=MONEY, color='000000')
        differ = abs(reel - hab) > 0.001
        c = dc(ws, row, 4, reel, bg=bg, fmt=MONEY,
               color=('C0392B' if reel > hab else ('1D9E75' if differ else '000000')),
               bold=differ)
        for ci in range(5, nc + 1):
            dc(ws, row, ci, '', bg=bg)
        row += 1

    # Sous-totaux par type
    for typ, bg in [('Yssingeaux', C['ys2']), ('Le Puy', C['py2']), ('Général', C['gn2'])]:
        dc(ws, row, 1, f'Sous-total {typ}', bg=bg, align='right',
           bold=True, color=type_bg[typ].replace(type_bg[typ], {C['ys']: C['ys3'], C['py']: C['py3'], C['gn']: C['gn3']}[type_bg[typ]]))
        dc(ws, row, 2, '', bg=bg); dc(ws, row, 3, '', bg=bg)
        dc(ws, row, 4, type_tot[typ], bg=bg, fmt=MONEY, bold=True)
        for ci in range(5, nc + 1): dc(ws, row, ci, '', bg=bg)
        row += 1

    # Total charges
    dc(ws, row, 1, 'TOTAL CHARGES', bg=C['dk'], color=C['wh'], bold=True, align='right')
    dc(ws, row, 2, '', bg=C['dk']); dc(ws, row, 3, tot_hab, bg=C['dk'], color=C['wh'], fmt=MONEY)
    dc(ws, row, 4, tot_reel, bg=C['dk'], color=C['wh'], fmt=MONEY, bold=True)
    for ci in range(5, nc + 1): dc(ws, row, ci, '', bg=C['dk'])
    row += 2

    # ── Section 3 : CA LE PUY & REDEVANCE ─────────────────────────────────────
    merge_hc(ws, row, 1, row, nc,
             f'3. CA LE PUY & REDEVANCE MATÉRIEL ({tx*100:.1f}%)', bg=C['md'], sz=10)
    row += 1
    hc(ws, row, 1, 'Orthoptiste', bg=C['lt'], fg=C['dk'], align='left')
    hc(ws, row, 2, '', bg=C['lt'])
    hc(ws, row, 3, 'CA Le Puy (€)', bg=C['py'], fg=C['py3'])
    hc(ws, row, 4, f'Redevance {tx*100:.1f}% (€)', bg=C['py'], fg=C['py3'])
    for ci in range(5, nc + 1): hc(ws, row, ci, '', bg=C['lt'])
    row += 1
    tot_ca = 0; tot_rev = 0
    for a in assocs:
        ca = get_ca(data, yr, mo, a['id'])
        rev = ca * tx
        tot_ca += ca; tot_rev += rev
        dc(ws, row, 1, a['nom'], bg=C['py'], align='left')
        dc(ws, row, 2, '', bg=C['py'])
        dc(ws, row, 3, ca, bg=C['py'], fmt=MONEY, color=C['py3'])
        dc(ws, row, 4, rev, bg=C['py'], fmt=MONEY, bold=True, color=C['py3'])
        for ci in range(5, nc + 1): dc(ws, row, ci, '', bg=C['py'])
        row += 1
    dc(ws, row, 1, 'TOTAL', bg=C['dk'], color=C['wh'], bold=True, align='right')
    dc(ws, row, 2, '', bg=C['dk'])
    dc(ws, row, 3, tot_ca, bg=C['dk'], color=C['wh'], fmt=MONEY, bold=True)
    dc(ws, row, 4, tot_rev, bg=C['dk'], color=C['wh'], fmt=MONEY, bold=True)
    for ci in range(5, nc + 1): dc(ws, row, ci, '', bg=C['dk'])
    row += 2

    # ── Section 4 : RÉPARTITION ────────────────────────────────────────────────
    merge_hc(ws, row, 1, row, nc, '4. RÉPARTITION PAR ORTHOPTISTE', bg=C['md'], sz=10)
    row += 1
    headers4 = ['Poste', ''] + [a['nom'] for a in assocs] + ['TOTAL']
    for ci, h in enumerate(headers4, 1):
        hc(ws, row, ci, h, bg=C['lt'], fg=C['dk'], wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1

    rep = calc_repartition(data, yr, mo)
    parts = [
        ('Part Yssingeaux',    'y',     C['ys'],  C['ys3']),
        ('Part Le Puy',        'p',     C['py'],  C['py3']),
        ('Part Général',       'g',     C['gn'],  C['gn3']),
        ('Redevance matériel', 'rev',   C['rev'], '7A5A00'),
        ('Total dû',           'tot',   C['tot'], C['md']),
        ('Avance versée',      'av',    C['av'],  '4A235A'),
        ('Reste à payer',      'reste', C['dk'],  C['wh']),
    ]
    for lbl, key, bg, fg in parts:
        bold = key in ('tot', 'reste')
        dc(ws, row, 1, lbl, bg=bg, align='left', color=fg, bold=bold)
        dc(ws, row, 2, '', bg=bg)
        row_tot = 0
        for i, a in enumerate(assocs):
            val = rep.get(a['id'], {}).get(key, 0)
            row_tot += val
            c = dc(ws, row, 3 + i, val, bg=bg, color=fg, fmt=MONEY, bold=bold)
            if key == 'reste' and val < 0.01:
                c.font = Font(name=FONT, color='1D9E75', bold=True, size=10)
        dc(ws, row, 3 + NA, row_tot, bg=bg, color=fg, fmt=MONEY, bold=bold)
        row += 1

    ws.freeze_panes = 'A4'
    return ws

# ── ONGLET PAR ORTHOPTISTE ─────────────────────────────────────────────────────
def build_ortho(wb, data, yr, assoc):
    ws = wb.create_sheet(assoc['nom'][:20])
    ws.sheet_view.showGridLines = False
    aid = assoc['id']
    tx = taux_redev(data)

    ws.column_dimensions['A'].width = 26
    for i in range(2, 16):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.row_dimensions[1].height = 30
    nc = 14
    merge_hc(ws, 1, 1, 1, nc,
             f'SCM Vidi — {assoc["nom"]} — Année {yr}', sz=13)
    ws.row_dimensions[2].height = 18
    merge_hc(ws, 2, 1, 2, nc,
             f'Taux redevance matériel : {tx*100:.1f}%', bg=C['md'], sz=9, bold=False)

    # En-têtes mois
    ws.row_dimensions[4].height = 28
    hc(ws, 4, 1, 'Indicateur', bg=C['lt'], fg=C['dk'], align='left')
    for mi, m in enumerate(MOIS):
        hc(ws, 4, 2 + mi, m, bg=C['lt'], fg=C['dk'], wrap=True)
    hc(ws, 4, 14, 'TOTAL', bg=C['dk'])

    row = 5
    sections = [
        ('DJ Yssingeaux',     lambda mo: dj_count(data, yr, mo, 'Yssingeaux', aid), C['ys'],  C['ys3'], '0'),
        ('DJ Le Puy',         lambda mo: dj_count(data, yr, mo, 'Le Puy', aid),     C['py'],  C['py3'], '0'),
        ('Part Yssingeaux',   lambda mo: calc_repartition(data, yr, mo).get(aid, {}).get('y', 0),     C['ys'],  C['ys3'], MONEY),
        ('Part Le Puy',       lambda mo: calc_repartition(data, yr, mo).get(aid, {}).get('p', 0),     C['py'],  C['py3'], MONEY),
        ('Part Général',      lambda mo: calc_repartition(data, yr, mo).get(aid, {}).get('g', 0),     C['gn'],  C['gn3'], MONEY),
        ('CA Le Puy',         lambda mo: get_ca(data, yr, mo, aid),                                    C['py'],  C['py3'], MONEY),
        ('Redevance matériel',lambda mo: get_ca(data, yr, mo, aid) * tx,                               C['rev'], '7A5A00', MONEY),
        ('Total dû',          lambda mo: calc_repartition(data, yr, mo).get(aid, {}).get('tot', 0),    C['tot'], C['md'],  MONEY),
        ('Avance versée',     lambda mo: get_avance(data, yr, mo, aid),                                C['av'],  '4A235A', MONEY),
        ('Reste à payer',     lambda mo: calc_repartition(data, yr, mo).get(aid, {}).get('reste', 0),  C['dk'],  C['wh'],  MONEY),
    ]

    for lbl, fn, bg, fg, fmt in sections:
        bold = lbl in ('Total dû', 'Reste à payer')
        dc(ws, row, 1, lbl, bg=bg, align='left', color=fg, bold=bold)
        annual = 0
        for mo in range(12):
            val = fn(mo)
            annual += val
            c = dc(ws, row, 2 + mo, val, bg=bg, color=fg, fmt=fmt, bold=bold)
            if lbl == 'Reste à payer' and val < 0.01:
                c.font = Font(name=FONT, color='1D9E75', bold=True, size=10)
        dc(ws, row, 14, annual, bg=bg, color=fg, fmt=fmt, bold=bold)
        row += 1
        if lbl in ('DJ Le Puy', 'Part Général', 'Redevance matériel'):
            for ci in range(1, 15):
                ws.cell(row=row, column=ci).fill = PatternFill('solid', fgColor='F0EDE8')
            row += 1

    ws.freeze_panes = 'B5'
    return ws

# ── MAIN ───────────────────────────────────────────────────────────────────────
def main():
    # Chercher le fichier JSON
    json_file = None
    annee = None

    for arg in sys.argv[1:]:
        if arg.endswith('.json'):
            json_file = Path(arg)
        elif arg.isdigit():
            annee = int(arg)

    if json_file is None:
        # Chercher automatiquement dans le dossier courant
        candidates = sorted(Path('.').glob('scm_vidi*.json'))
        if not candidates:
            print("❌ Aucun fichier scm_vidi*.json trouvé.")
            print("   Usage : python export_excel.py scm_vidi_data.json 2025")
            sys.exit(1)
        json_file = candidates[-1]
        print(f"  Fichier JSON : {json_file}")

    data = load(json_file)

    # Déterminer l'année
    if annee is None:
        # Regarder dans les données
        annees = set()
        for section in ['plannings', 'montantsReels', 'caPuy', 'avances']:
            annees.update(int(k) for k in data.get(section, {}).keys() if k.isdigit())
        annee = max(annees) if annees else date.today().year
        print(f"  Année détectée : {annee}")

    assocs = [a for a in data['associes'] if a.get('actif')]
    print(f"  Associés : {', '.join(a['nom'] for a in assocs)}")
    print(f"  Taux redevance : {taux_redev(data)*100:.1f}%")
    print(f"  Construction du fichier Excel…")

    wb = Workbook()

    # 1. Récap annuel (onglet actif)
    build_recap(wb, data, annee)

    # 2. Onglets mensuels
    for mo in range(12):
        build_mois(wb, data, annee, mo)
        print(f"    ✓ {MOIS[mo]}")

    # 3. Onglets par orthoptiste
    for a in assocs:
        build_ortho(wb, data, annee, a)
        print(f"    ✓ {a['nom']}")

    out = Path(f'SCM_Vidi_{annee}.xlsx')
    wb.save(out)
    print(f"\n  Recalcul des formules…")

    # Recalcul
    import subprocess, json as jmod
    result = subprocess.run(
        ['python', str(Path(__file__).parent / 'scripts' / 'recalc.py'), str(out), '60'],
        capture_output=True, text=True)
    try:
        info = jmod.loads(result.stdout)
        errs = info.get('total_errors', 0)
        fmls = info.get('total_formulas', 0)
        if errs == 0:
            print(f"  ✓ {fmls} formules, 0 erreur")
        else:
            print(f"  ⚠ {errs} erreur(s) — {info.get('error_summary')}")
    except Exception:
        pass  # recalc optionnel

    print(f"\n✅ Fichier créé : {out.resolve()}")

if __name__ == '__main__':
    main()
